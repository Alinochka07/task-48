from flask import Flask, render_template, redirect, request
from openpyxl import Workbook, load_workbook


app = Flask(__name__)

@app.route('/')
def homepage():
    try:
        excel = load_workbook('goods.xlsx')
    except:
        excel = Workbook()
    
    page = excel[excel.sheetnames[0]]
    # goods = [cell.value for row in page for cell in row]
    goods = [row[0].value for row in page]
    excel.save('goods.xlsx')
    return render_template('index.html', goods=goods)


@app.route('/add/', methods=['POST'])
def add_goods():
    good = request.form['good']
    try:
        excel = load_workbook('goods.xlsx')
    except:
        excel = Workbook()

    page = excel[excel.sheetnames[0]]
    end = len(page['A']) + 1
    page[f'A{end}'] = good
    excel.save('goods.xlsx')
     
    return """
            <h1>Товары добавлены</h1>
            <a href='/'>Вернуться на главную страницу</a>
            """ 
 
       
